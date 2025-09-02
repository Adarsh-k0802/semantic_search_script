import os
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from sentence_transformers import SentenceTransformer
from dotenv import load_dotenv
import json
import re
import google.generativeai as genai
from typing import List, Dict, Any, Tuple
import numpy as np
from pathlib import Path
import faiss

# --- Configuration ---
load_dotenv()

# Load API keys from environment variables
google_api_key = os.getenv('GOOGLE_API_KEY')
if not google_api_key:
    raise ValueError("Please set the GOOGLE_API_KEY environment variable in your .env file")

# Configure Gemini
genai.configure(api_key=google_api_key)

# Initialize components
embedder = SentenceTransformer('all-MiniLM-L6-v2')

# Global variables for FAISS
vector_dimension = 384
index = faiss.IndexFlatL2(vector_dimension)
stored_texts = []
stored_metadata = []
global_numeric_data = []  # Store numeric values for calculations

# Initialize Gemini model
try:
    gemini_model_instance = genai.GenerativeModel('gemini-1.5-flash')
except Exception as e:
    print(f"Error initializing Gemini model: {e}")
    gemini_model_instance = None


# --- Enhanced Semantic Keyword Knowledge Graph ---
semantic_keyword_graph = {
    
# FINANCIAL PERFORMANCE METRICS (Actual Results)
    "actual revenue": ["realized revenue", "achieved sales", "closed revenue", "earned income", "final sales figure", "confirmed revenue"],
    "actual sales": ["achieved sales", "completed transactions", "realized deals", "finalized revenue", "closed deals value"],
    "net profit": ["bottom line", "net income", "final profit", "earnings after tax", "post-tax profit", "actual profit"],
    "gross profit": ["revenue minus cogs", "trading profit", "pre-operating profit", "direct margin", "manufacturing profit"],
    
    # FINANCIAL TARGETS/FORECASTS (Planned/Expected)
    "target revenue": ["planned revenue", "budgeted sales", "expected income", "revenue goal", "projected sales", "revenue forecast"],
    "forecast": ["projection", "outlook", "budget", "financial plan", "future estimate", "prediction", "3-year forecast"],
    "quota": ["sales target", "revenue goal", "individual target", "performance expectation", "sales objective", "assigned goal"],
    
    # FINANCIAL RATIOS & MARGINS
    "gross margin": ["gross profit ratio", "direct margin percentage", "revenue-cogs difference %", "trading margin", "production efficiency ratio"],
    "net profit margin": ["net income ratio", "bottom line percentage", "after-tax margin", "overall profitability ratio", "final margin %"],
    "operating margin": ["operating profit ratio", "ebit margin", "business efficiency ratio", "core operations margin"],
    "expense ratio": ["cost efficiency ratio", "overhead percentage", "operating cost ratio", "expense-to-revenue ratio"],
    "cogs ratio": ["cost of goods percentage", "direct cost ratio", "production cost percentage", "variable cost ratio"],
    
    # GROWTH METRICS
    "revenue growth": ["sales growth", "top line expansion", "income increase percentage", "year-over-year revenue change", "growth rate"],
    "net profit growth": ["bottom line growth", "earnings expansion", "profit increase percentage", "year-over-year profit change"],
    "growth %": ["performance increase", "improvement rate", "expansion percentage", "progress metric", "period-over-period change"],
    
    # COST & EXPENSE CATEGORIES
    "cost of goods sold": ["cogs", "direct costs", "production costs", "variable costs", "manufacturing expenses", "cost of sales"],
    "operating expenses": ["opex", "overhead costs", "indirect costs", "fixed costs", "administrative expenses", "business expenses"],
    "salaries": ["wages", "compensation", "payroll", "employee costs", "staff expenses", "labor costs"],
    "marketing": ["advertising", "promotion", "customer acquisition costs", "brand expenses", "sales support costs"],
    
    # COMPARATIVE & VARIANCE METRICS
    "variance": ["performance gap", "deviation from target", "actual vs planned difference", "delta", "shortfall/excess", "budget variance"],
    "performance %": ["achievement rate", "quota attainment", "target completion percentage", "goal fulfillment rate", "actual vs target ratio"],
    "% to target": ["goal attainment percentage", "target achievement rate", "quota completion %", "performance to plan ratio"],
    
    # TIME PERIODS & REPORTING
    "year 1": ["first year", "initial period", "base year", "year one", "starting fiscal year"],
    "year 2": ["second year", "follow-on year", "intermediate period", "year two", "growth year"],
    "year 3": ["third year", "final forecast year", "mature period", "year three", "projection end year"],
    "monthly cost": ["per month expense", "monthly expenditure", "recurring monthly expense", "fixed monthly cost"],
    "annual cost": ["yearly expense", "annual expenditure", "total yearly cost", "12-month expense"],
    
    # SALES & PIPELINE CONCEPTS (From Sales Dashboard)
    "pipeline": ["sales funnel", "opportunity pipeline", "deal progression", "sales process", "prospect pipeline"],
    "deal stage": ["sales phase", "pipeline stage", "prospecting", "qualified", "proposal", "negotiation", "closed won", "closed lost"],
    "probability %": ["win likelihood", "closure chance", "deal confidence", "conversion probability", "expected closure rate"],
    
    # CUSTOMER ANALYSIS CONCEPTS (From Sales Dashboard)
    "segment": ["customer classification", "market segment", "business type", "enterprise", "smb", "customer category"],
    "churned": ["lost customer", "customer attrition", "defected account", "canceled service", "discontinued business"],
    "last purchase": ["most recent transaction", "final order date", "latest sale timestamp", "previous acquisition date"],
    
    # PRODUCT CONCEPTS (From Sales Dashboard)
    "units sold": ["quantity moved", "sales volume", "items purchased", "transaction count", "product movement"],
    "category": ["product type", "classification group", "hardware", "software", "service", "product family"],
    
    # TEMPORAL CONCEPTS
    "month": ["monthly period", "calendar month", "30-day cycle", "billing period", "reporting month"],
    "close date": ["expected completion", "deal finalization date", "anticipated closure", "projected end date"],
    
    # GEOGRAPHICAL CONCEPTS
    "region": ["sales territory", "geographical area", "district", "zone", "north", "south", "east", "west"],
    
    # ORGANIZATIONAL CONCEPTS
    "rep": ["sales representative", "account executive", "salesperson", "sales agent", "account manager"],
    "customer": ["client", "account", "business", "organization", "company", "buyer"],

    
    # Half-year identifiers
    "h1": ["first half", "h1", "half 1", "jan-jun", "january-june"],
    "h2": ["second half", "h2", "half 2", "jul-dec", "july-december"],
    
    # Months - Full names and abbreviations
    "january": ["jan", "january", "month 1", "first month"],
    "february": ["feb", "february", "month 2", "second month"],
    "march": ["mar", "march", "month 3", "third month"],
    "april": ["apr", "april", "month 4", "fourth month"],
    "may": ["may", "month 5", "fifth month"],
    "june": ["jun", "june", "month 6", "sixth month"],
    "july": ["jul", "july", "month 7", "seventh month"],
    "august": ["aug", "august", "month 8", "eighth month"],
    "september": ["sep", "sept", "september", "month 9", "ninth month"],
    "october": ["oct", "october", "month 10", "tenth month"],
    "november": ["nov", "november", "month 11", "eleventh month"],
    "december": ["dec", "december", "month 12", "twelfth month", "year end", "yearend"],
    
    # Date Formats and Patterns (generic patterns)
    "date": ["date", "timestamp", "period", "timeframe", "duration", "interval"],
    "year_format": ["yyyy", "year", "yy"],
    "month_format": ["mm", "month", "mon"],
    "day_format": ["dd", "day"],
    "date_separator": ["-", "/", ".", " "],
    
    # Time-related concepts
    "current": ["current", "present", "now", "latest", "most recent", "this period"],
    "previous": ["previous", "last", "prior", "preceding", "former", "past"],
    "next": ["next", "upcoming", "future", "forthcoming", "subsequent", "following"],
    "historical": ["historical", "past", "archive", "record", "historical data"],
    
    # Year context (without specific years)
    "current_year": ["current year", "this year", "present year"],
    "previous_year": ["last year", "previous year", "prior year", "year before"],
    "next_year": ["next year", "upcoming year", "following year"],
    "fiscal_year": ["fiscal year", "financial year", "accounting year"],
    
    # Seasonal periods
    "spring": ["spring", "q2", "growing season"],
    "summer": ["summer", "q3", "peak season"],
    "fall": ["fall", "autumn", "q4", "harvest season"],
    "winter": ["winter", "q1", "holiday season"],
    
    # Comparison terms
    "vs": ["vs", "versus", "compared to", "against", "relative to", "comparison"],
    "change": ["change", "delta", "difference", "variance", "deviation", "shift"],
    "increase": ["increase", "growth", "rise", "up", "gain", "improvement"],
    "decrease": ["decrease", "decline", "drop", "down", "loss", "reduction"],
    "comparison": ["compare", "comparison", "versus", "vs", "relative", "benchmark"],
    
    # Measurement units
    "currency": ["$", "dollar", "usd", "currency", "money", "cash", "amount", "value", "financial"],
    "percentage": ["%", "percent", "percentage", "rate", "ratio", "pct"],
    "unit": ["units", "quantity", "volume", "count", "number", "items"],
    
    # Data quality terms
    "actual": ["actual", "real", "realized", "achieved", "actuals", "result"],
    "target": ["target", "goal", "budget", "forecast", "projection", "plan", "estimate"],
    "variance": ["variance", "difference", "gap", "deviation", "discrepancy", "variation"],
    "budget": ["budget", "planned", "estimated", "projected", "forecasted"],
    
    # Time range indicators
    "range": ["range", "period", "timeframe", "duration", "from-to", "between"],
    "start_date": ["start", "beginning", "from", "since", "effective date"],
    "end_date": ["end", "until", "to", "through", "expiration"],
    
    # Frequency indicators
    "daily": ["daily", "day", "per day", "each day"],
    "weekly": ["weekly", "week", "per week", "each week"],
    "monthly": ["monthly", "month", "per month", "each month"],
    "quarterly": ["quarterly", "quarter", "per quarter", "each quarter"],
    "yearly": ["yearly", "annual", "per year", "each year", "annually"],
    
    # Relative time indicators
    "recent": ["recent", "latest", "current", "present", "now"],
    "historical": ["historical", "past", "previous", "prior", "old"],
    "future": ["future", "upcoming", "next", "forthcoming", "planned"],
    
    # Date component recognition
    "year_component": ["year", "yr", "yyyy", "yy"],
    "month_component": ["month", "mth", "mm", "mon"],
    "day_component": ["day", "dd", "date"],
    "week_component": ["week", "wk", "ww"],
    
    # Special time periods
    "ytd": ["ytd", "year to date", "year-to-date"],
    "mtd": ["mtd", "month to date", "month-to-date"],
    "qtd": ["qtd", "quarter to date", "quarter-to-date"],
    "rolling": ["rolling", "moving", "trailing", "sliding"],
    
    # Time aggregation
    "total": ["total", "sum", "aggregate", "overall", "cumulative"],
    "average": ["average", "avg", "mean", "typical", "normal"],
    "maximum": ["maximum", "max", "highest", "peak", "top"],
    "minimum": ["minimum", "min", "lowest", "bottom", "trough"],
    
     # Date Format Patterns and Identifiers
    "date_formats": [
        # International formats
        "dd/mm/yyyy", "dd-mm-yyyy", "dd.mm.yyyy", "dd mm yyyy",
        "dd/mmm/yyyy", "dd-mmm-yyyy", "dd.mmm.yyyy", "dd mmm yyyy",
        "dd/mm/yy", "dd-mm-yy", "dd.mm.yy", "dd mm yy",
        
        # US formats
        "mm/dd/yyyy", "mm-dd-yyyy", "mm.dd.yyyy", "mm dd yyyy", 
        "mmm/dd/yyyy", "mmm-dd-yyyy", "mmm.dd.yyyy", "mmm dd yyyy",
        "mm/dd/yy", "mm-dd-yy", "mm.dd.yy", "mm dd yy",
        
        # ISO and standard formats
        "yyyy-mm-dd", "yyyy/mm/dd", "yyyy.mm.dd", "yyyy mm dd",
        "yyyy-mmm-dd", "yyyy/mmm/dd", "yyyy.mmm.dd", "yyyy mmm dd",
        "yy-mm-dd", "yy/mm/dd", "yy.mm.dd", "yy mm dd",
        
        # Month-year formats
        "mmm yyyy", "mmmm yyyy", "mm-yyyy", "mm/yyyy", "mm.yyyy", "mm yyyy",
        "yyyy mmm", "yyyy mmmm", "yyyy-mm", "yyyy/mm", "yyyy.mm", "yyyy mm",
        
        # Day-month formats
        "dd mmm", "dd mmmm", "dd-mm", "dd/mm", "dd.mm", "dd mm",
        
        # Compact formats
        "ddmmyyyy", "ddmmyy", "mmddyyyy", "mmddyy", "yyyymmdd",
        
        # Written formats
        "day month year", "month day year", "year month day",
        "day month", "month day", "month year", "year month"
    ],
    
    # Date Component Patterns
    "day_patterns": [
        "dd", "d", "day", "date", 
        # Day numbers with various padding
        "01", "02", "03", "04", "05", "06", "07", "08", "09", "10",
        "11", "12", "13", "14", "15", "16", "17", "18", "19", "20",
        "21", "22", "23", "24", "25", "26", "27", "28", "29", "30", "31"
    ],
    
    "month_patterns": [
        # Numeric months
        "mm", "m", "month", "mon", 
        "01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12",
        # Month abbreviations
        "jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec",
        # Full month names
        "january", "february", "march", "april", "may", "june", 
        "july", "august", "september", "october", "november", "december"
    ],
    
    "year_patterns": [
        "yyyy", "yy", "year", "yr",
        # Common year patterns (without specific years)
        "20", "19", "21", "22", "23", "24", "25", "26", "27", "28", "29",
        # Two-digit year patterns
        "00", "01", "02", "03", "04", "05", "06", "07", "08", "09",
        "10", "11", "12", "13", "14", "15", "16", "17", "18", "19",
        "20", "21", "22", "23", "24", "25", "26", "27", "28", "29",
        "30", "31", "32", "33", "34", "35", "36", "37", "38", "39"
    ],
    
    # Date Separators
    "date_separators": ["/", "-", ".", " ", ":", "|", "\\", ",", ""],
    
    # Date Format Type Identifiers
    "european_format": ["dd/mm/yyyy", "dd-mm-yyyy", "dd.mm.yyyy", "dd mm yyyy", "ddmmyyyy"],
    "us_format": ["mm/dd/yyyy", "mm-dd-yyyy", "mm.dd.yyyy", "mm dd yyyy", "mmddyyyy"],
    "iso_format": ["yyyy-mm-dd", "yyyy/mm/dd", "yyyy.mm.dd", "yyyy mm dd", "yyyymmdd"],
    "month_first": ["mm/", "mm-", "mm.", "mm ", "mmm/", "mmm-", "mmm.", "mmm "],
    "day_first": ["dd/", "dd-", "dd.", "dd ", "ddd/", "ddd-", "ddd.", "ddd "],
    "year_first": ["yyyy/", "yyyy-", "yyyy.", "yyyy ", "yy/", "yy-", "yy.", "yy "],
    
    # Date Validation Patterns
    "valid_date_indicators": [
        "date", "timestamp", "day", "month", "year", "period",
        "created", "modified", "updated", "effective", "expiry",
        "start", "end", "from", "to", "since", "until"
    ],
    
    # Date Context Keywords
    "date_context": [
        "on", "at", "in", "during", "for", "from", "to", "between",
        "after", "before", "since", "until", "by", "within"
    ],
    
    # Special Date Patterns
    "quarter_dates": ["q1", "q2", "q3", "q4", "quarter", "qtr"],
    "half_year_dates": ["h1", "h2", "half", "semester"],
    "week_patterns": ["week", "wk", "w", "weekly"],
    "season_patterns": ["spring", "summer", "fall", "autumn", "winter"],
    
    # Date Range Patterns
    "date_range": [
        "to", "until", "through", "-", "/", "~", "...", "..",
        "from", "between", "and", "&"
    ],
    
    # Time Components (for datetime recognition)
    "time_patterns": [
        "hh:mm:ss", "hh:mm", "hhmm", "h:mm", "h.mm",
        "am", "pm", "a.m.", "p.m.", "noon", "midnight"
    ],
    
    # Locale-specific date indicators
    "european_date_indicators": [
        "european", "euro", "uk", "british", "french", "german", "italian", "spanish",
        "australian", "new zealand", "australia", "nz", "commonwealth"
    ],
    
    "us_date_indicators": [
        "us", "american", "usa", "united states", "north american"
    ],
    
    # Date parsing helpers
    "date_parsing_keywords": [
        "format", "pattern", "style", "type", "layout",
        "dd", "mm", "yyyy", "yy", "month", "day", "year",
        "separator", "delimiter", "order"
    ]
}

def detect_date_format(text):
    """Detect date format patterns in text."""
    if not text or not isinstance(text, str):
        return None
    
    text_lower = text.lower()
    
    # Check for specific date format patterns
    for format_pattern in semantic_keyword_graph["date_formats"]:
        if format_pattern in text_lower:
            return format_pattern
    
    # Check for component patterns
    has_day = any(pattern in text_lower for pattern in semantic_keyword_graph["day_patterns"])
    has_month = any(pattern in text_lower for pattern in semantic_keyword_graph["month_patterns"])
    has_year = any(pattern in text_lower for pattern in semantic_keyword_graph["year_patterns"])
    
    if has_day and has_month and has_year:
        # Try to determine order
        if any(indicator in text_lower for indicator in semantic_keyword_graph["day_first"]):
            return "dd/mm/yyyy"
        elif any(indicator in text_lower for indicator in semantic_keyword_graph["month_first"]):
            return "mm/dd/yyyy"
        elif any(indicator in text_lower for indicator in semantic_keyword_graph["year_first"]):
            return "yyyy-mm-dd"
        return "date_format_detected"
    
    return None

def is_likely_date_header(header_text):
    """Check if header text is likely a date column."""
    if not header_text or not isinstance(header_text, str):
        return False
    
    text_lower = header_text.lower()
    
    # Check for date context keywords
    if any(keyword in text_lower for keyword in semantic_keyword_graph["valid_date_indicators"]):
        return True
    
    # Check for date format patterns
    if detect_date_format(text_lower):
        return True
    
    # Check for date component patterns
    day_patterns = any(pattern in text_lower for pattern in semantic_keyword_graph["day_patterns"])
    month_patterns = any(pattern in text_lower for pattern in semantic_keyword_graph["month_patterns"])
    year_patterns = any(pattern in text_lower for pattern in semantic_keyword_graph["year_patterns"])
    
    return day_patterns or month_patterns or year_patterns

def parse_date_cell(value, format_hint=None):
    """Parse date cell value with format detection."""
    if not value:
        return None
    
    if isinstance(value, (int, float)):
        # Excel serial date number
        try:
            from datetime import datetime, timedelta
            excel_base_date = datetime(1899, 12, 30)
            return excel_base_date + timedelta(days=value)
        except:
            return None
    
    if isinstance(value, str):
        # String date - try to parse with various formats
        try:
            from dateutil import parser
            return parser.parse(value, fuzzy=True)
        except:
            return None
    
    return None

# --- Advanced Header Detection Functions ---
def detect_table_structure(sheet):
    """Detect table structure assuming only one table per sheet."""
    # Find the first row with data
    start_row = None
    for row_idx in range(1, sheet.max_row + 1):
        if any(sheet.cell(row=row_idx, column=col_idx).value is not None 
              for col_idx in range(1, sheet.max_column + 1)):
            start_row = row_idx
            break
    
    if start_row is None:
        return []  # No data found
    
    # Find the last row with data
    end_row = start_row
    for row_idx in range(start_row + 1, sheet.max_row + 1):
        if any(sheet.cell(row=row_idx, column=col_idx).value is not None 
              for col_idx in range(1, sheet.max_column + 1)):
            end_row = row_idx
    
    # Find the first column with data
    start_col = None
    for col_idx in range(1, sheet.max_column + 1):
        if any(sheet.cell(row=row_idx, column=col_idx).value is not None 
              for row_idx in range(start_row, end_row + 1)):
            start_col = col_idx
            break
    
    if start_col is None:
        return []  # No data found
    
    # Find the last column with data
    end_col = start_col
    for col_idx in range(start_col + 1, sheet.max_column + 1):
        if any(sheet.cell(row=row_idx, column=col_idx).value is not None 
              for row_idx in range(start_row, end_row + 1)):
            end_col = col_idx
    
    # Create a single table
    table = {
        'start_row': start_row,
        'end_row': end_row,
        'start_col': start_col,
        'end_col': end_col,
        'headers': {}
    }
    
    # Detect headers (assume first row is headers)
    header_row = start_row
    for col_idx in range(start_col, end_col + 1):
        header_cell = sheet.cell(row=header_row, column=col_idx)
        if header_cell.value:
            table['headers'][col_idx] = str(header_cell.value)
    
    return [table]

def get_cell_headers(sheet, row_idx, col_idx, tables):
    """Get proper row and column headers for a cell with date format detection."""
    row_header = None
    column_header = None
    date_format = None
    
    # Find which table this cell belongs to
    for table in tables:
        if table['start_row'] <= row_idx <= table['end_row'] and table['start_col'] <= col_idx <= table['end_col']:
            # Column header from table headers
            if col_idx in table['headers']:
                column_header = table['headers'][col_idx]
                # Detect date format in column header
                date_format = detect_date_format(column_header)
            
            # Row header - look for first column in this row within table
            for test_col in range(table['start_col'], col_idx):
                test_cell = sheet.cell(row=row_idx, column=test_col)
                if test_cell.value and str(test_cell.value).strip():
                    row_header = str(test_cell.value).strip()
                    # Detect date format in row header
                    if not date_format:
                        date_format = detect_date_format(row_header)
                    break
            
            break
    
    return row_header, column_header, date_format

def extract_cell_references_simple(formula):
    """Simple and reliable formula reference extraction."""
    if not formula or not formula.startswith('='):
        return []
    
    # Simple pattern matching for cell references
    patterns = [
        r"'([^']+)'!([A-Z]+\d+)",  # 'Sheet'!A1
        r"([A-Z]+\d+)",  # Simple A1
    ]
    
    references = []
    for pattern in patterns:
        matches = re.findall(pattern, formula)
        for match in matches:
            if isinstance(match, tuple):
                references.append(f"'{match[0]}'!{match[1]}")
            else:
                references.append(match)
    
    return list(set(references))

def create_semantic_description(sheet_name, cell_address, value, formula, row_header, column_header, dependencies):
    """Create clear semantic description for a cell."""
    parts = []
    
    # Basic identification
    parts.append(f"In sheet '{sheet_name}', cell {cell_address}")
    
    # Add headers context
    if row_header and column_header:
        parts.append(f"represents '{row_header} - {column_header}'.")
    elif row_header:
        parts.append(f"represents '{row_header}'.")
    elif column_header:
        parts.append(f"represents '{column_header}'.")
    else:
        parts.append("contains important data.")
    
    # Add value
    if value is not None:
        parts.append(f"Value: {value}")
    
    # Add formula if exists
    if formula:
        parts.append(f"Formula: {formula}")
    
    # Add dependencies if any
    if dependencies:
        parts.append("Depends on: " + ", ".join(dependencies[:3]))
        if len(dependencies) > 3:
            parts.append(f"(and {len(dependencies) - 3} more)")
    
    return " ".join(parts)

# --- Main Parsing Function ---
def parse_spreadsheet_structured(file_path):
    """Parse spreadsheet with proper table structure understanding and numeric data."""
    wb = load_workbook(filename=file_path, data_only=False)
    data_wb = load_workbook(filename=file_path, data_only=True)
    
    chunks = []
    all_metadata = []
    numeric_data = []  # Store numeric values for calculations
    
    print("Starting structured spreadsheet parsing...")
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        data_sheet = data_wb[sheet_name]
        
        print(f"📊 Processing sheet: {sheet_name}")
        
        # Detect table structure
        tables = detect_table_structure(sheet)
        print(f"   Found {len(tables)} tables")
        
        for table in tables:
            print(f"   Processing table from row {table['start_row']} to {table['end_row']}")
            
            for row_idx in range(table['start_row'] + 1, table['end_row'] + 1):
                for col_idx in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    
                    if cell.value is None:
                        continue
                    
                    data_cell = data_sheet.cell(row=row_idx, column=col_idx)
                    value = data_cell.value
                    formula = str(cell.value) if str(cell.value).startswith('=') else None
                    
                    # Get proper headers
                    # Get proper headers
                    row_header, column_header, date_format = get_cell_headers(sheet, row_idx, col_idx, [table])
                    
                    # Extract dependencies
                    dependencies = extract_cell_references_simple(formula) if formula else []
                    
                    # Create semantic description
                    semantic_text = create_semantic_description(
                        sheet_name, cell.coordinate, value, formula, 
                        row_header, column_header, dependencies
                    )
                    
                    # Store numeric data for calculations
                    numeric_info = {
                        "sheet": sheet_name,
                        "cell": cell.coordinate,
                        "value": value,
                        "row_header": row_header,
                        "column_header": column_header,
                        "is_numeric": isinstance(value, (int, float)) and value is not None
                    }
                    
                    if numeric_info["is_numeric"]:
                        numeric_data.append(numeric_info)
                    
                    # Create chunk
                    chunk = {
                        "text": semantic_text,
                        "metadata": {
                            "source_sheet": sheet_name,
                            "cell_address": cell.coordinate,
                            "value": str(value) if value is not None else "",
                            "numeric_value": value if isinstance(value, (int, float)) else None,
                            "formula": formula,
                            "row_header": row_header,
                            "column_header": column_header,
                            "date_format": date_format,  # Add date format info
                            "dependencies": dependencies,
                            "is_numeric": isinstance(value, (int, float)) and value is not None
                        }
                    }
                    
                    chunks.append(chunk)
                    all_metadata.append(chunk["metadata"])
        
        print(f"   Processed {len([c for c in chunks if c['metadata']['source_sheet'] == sheet_name])} cells")
    
    # Store numeric data globally for calculations
    global global_numeric_data
    global_numeric_data = numeric_data
    
    print(f"✅ Total chunks created: {len(chunks)}")
    print(f"✅ Numeric cells found: {len(numeric_data)}")
    return chunks, all_metadata, numeric_data

# --- Enhanced Intent Detection System ---
def detect_query_intent(user_query):
    """Advanced intent detection using semantic patterns and LLM fallback."""
    query_lower = user_query.lower()
    
    # Semantic patterns for different intents
    intent_patterns = {
        "LOOKUP": [
            # Direct value requests
            r'find.*(value|number|amount|data|cell)',
            r'what.*(is|was).*\d{4}',
            r'show me.*(value|number|amount)',
            r'locate.*(cell|data|value)',
            r'where.*(is|are).*\d',
            # Specific value patterns
            r'target.*revenue.*jan',
            r'jan.*target.*revenue',
            r'revenue.*target.*jan',
            # General lookup patterns
            r'^what (is|was)',
            r'^find the',
            r'^show me the',
            r'^where is the'
        ],
        "CALCULATION": [
            # Math operations
            r'(sum|total|add|plus|combine).*of',
            r'average.*of',
            r'calculate.*(sum|total|average|mean)',
            r'how much.*total',
            r'what.*total.*of',
            # Statistical operations
            r'min.*of|max.*of',
            r'range.*of',
            r'standard deviation',
            # Aggregation patterns
            r'all.*revenue',
            r'total.*sales',
            r'combined.*profit'
        ],
        "EXPLANATION": [
            r'how.*work',
            r'explain.*formula',
            r'what does.*calculate',
            r'how is.*calculated',
            r'meaning of.*formula'
        ],
        "COMPARISON": [
            r'compare.*to',
            r'vs\.|versus',
            r'difference between',
            r'better than|worse than',
            r'higher than|lower than'
        ]
    }
    
    # Score each intent based on pattern matching
    intent_scores = {intent: 0 for intent in intent_patterns.keys()}
    
    for intent, patterns in intent_patterns.items():
        for pattern in patterns:
            if re.search(pattern, query_lower):
                intent_scores[intent] += 1
    
    # Add weight based on keywords (weaker evidence than patterns)
    keyword_weights = {
        "LOOKUP": ['find', 'locate', 'where', 'show', 'display', 'get'],
        "CALCULATION": ['calculate', 'compute', 'sum', 'total', 'average', 'add'],
        "EXPLANATION": ['explain', 'how', 'why', 'work', 'formula'],
        "COMPARISON": ['compare', 'versus', 'vs', 'difference', 'than']
    }
    
    for intent, keywords in keyword_weights.items():
        for keyword in keywords:
            if keyword in query_lower:
                intent_scores[intent] += 0.5
    
    # Get the highest scoring intent
    best_intent = max(intent_scores.items(), key=lambda x: x[1])
    # Use LLM as fallback for ambiguous cases or low confidence
    if best_intent[1] < 1.0:  # Low confidence score
        return llm_intent_detection(user_query)
    
    return best_intent[0]

def llm_intent_detection(query):
    """LLM fallback for ambiguous intent detection."""
    prompt = f"""
    Analyze this spreadsheet query and determine the primary intent:
    
    QUERY: "{query}"
    
    Choose ONE of these intents:
    - LOOKUP: Finding specific values, cells, or data points
    - CALCULATION: Performing math operations (sum, average, etc.)
    - EXPLANATION: Understanding formulas or relationships
    - COMPARISON: Comparing values or metrics
    - SEARCH: General data discovery
    
    Respond with ONLY the intent name in uppercase.
    """
    
    try:
        intent = call_gemini(prompt, temperature=0.0)
        return intent.strip().upper()
    except:
        return "LOOKUP"  # Default to lookup

# --- Enhanced Calculation Logic ---
def perform_calculations(search_results, user_query):
    """Improved calculation logic that respects intent."""
    query_lower = user_query.lower()
    
    # Extract numeric values with context awareness
    numeric_values = []
    
    for i, meta in enumerate(search_results['metadatas']):
        if meta.get('is_numeric') and meta.get('numeric_value') is not None:
            try:
                numeric_value = float(meta['numeric_value'])
                if not pd.isna(numeric_value):
                    # Calculate relevance based on query context
                    relevance = calculate_calculation_relevance(meta, query_lower)
                    
                    if relevance > 0.3:  # Minimum relevance threshold
                        numeric_values.append({
                            'value': numeric_value,
                            'relevance': relevance,
                            'metadata': meta,
                            'sheet': meta['source_sheet'],
                            'cell': meta['cell_address'],
                            'row_header': meta.get('row_header', ''),
                            'column_header': meta.get('column_header', '')
                        })
            except (ValueError, TypeError):
                continue
    
    if not numeric_values:
        return None
    
    # Sort by relevance
    numeric_values.sort(key=lambda x: x['relevance'], reverse=True)
    
    # Determine calculation type based on query semantics
    calc_type = detect_calculation_type(user_query)
    
    # Perform the appropriate calculation
    return execute_calculation(numeric_values, calc_type)

def calculate_calculation_relevance(metadata, query_lower):
    """Calculate relevance for calculation purposes."""
    relevance = 0.5  # Base relevance
    
    # Headers context
    row_header = str(metadata.get('row_header', '')).lower()
    col_header = str(metadata.get('column_header', '')).lower()
    full_context = f"{row_header} {col_header}"
    
    # Boost for header matches with query
    query_terms = [term for term in query_lower.split() if len(term) > 3]
    for term in query_terms:
        if term in full_context:
            relevance += 0.2
    
    # Penalize mismatched contexts
    if 'target' in query_lower and 'actual' in full_context:
        relevance -= 0.3
    if 'actual' in query_lower and 'target' in full_context:
        relevance -= 0.3
    
    # Boost for appropriate value ranges based on context
    value = metadata.get('numeric_value')
    if value is not None:
        if 'revenue' in query_lower and 1000 <= abs(value) <= 1000000:
            relevance += 0.1
        elif 'percentage' in query_lower and 0 <= abs(value) <= 100:
            relevance += 0.1
    
    return max(0, min(relevance, 1.0))  # Keep between 0-1

def detect_calculation_type(query):
    """Detect what type of calculation to perform."""
    query_lower = query.lower()
    
    calculation_keywords = {
        'sum': ['sum', 'total', 'add', 'plus', '+', 'combined', 'aggregate'],
        'average': ['average', 'avg', 'mean', 'typical'],
        'count': ['count', 'number of', 'how many'],
        'min': ['min', 'minimum', 'lowest', 'smallest'],
        'max': ['max', 'maximum', 'highest', 'largest'],
        'range': ['range', 'difference between', 'spread'],
        'median': ['median', 'middle value'],
        'std_dev': ['standard deviation', 'std dev', 'variation']
    }
    
    for calc_type, keywords in calculation_keywords.items():
        if any(keyword in query_lower for keyword in keywords):
            return calc_type
    
    return 'average'  # Default to average

def execute_calculation(numeric_values, calc_type):
    """Execute the appropriate calculation."""
    values = [item['value'] for item in numeric_values]
    
    if calc_type == 'sum' and values:
        result = sum(values)
        operation = 'SUM'
    elif calc_type == 'average' and values:
        result = sum(values) / len(values)
        operation = 'AVERAGE'
    elif calc_type == 'count':
        result = len(values)
        operation = 'COUNT'
    elif calc_type == 'min' and values:
        result = min(values)
        operation = 'MIN'
    elif calc_type == 'max' and values:
        result = max(values)
        operation = 'MAX'
    elif calc_type == 'range' and len(values) >= 2:
        result = max(values) - min(values)
        operation = 'RANGE'
    elif calc_type == 'median' and values:
        sorted_vals = sorted(values)
        n = len(sorted_vals)
        result = sorted_vals[n//2] if n % 2 == 1 else (sorted_vals[n//2-1] + sorted_vals[n//2]) / 2
        operation = 'MEDIAN'
    elif calc_type == 'std_dev' and len(values) >= 2:
        mean = sum(values) / len(values)
        variance = sum((x - mean) ** 2 for x in values) / len(values)
        result = variance ** 0.5
        operation = 'STANDARD_DEVIATION'
    else:
        return None
    
    return {
        'operation': operation,
        'result': result,
        'count': len(values),
        'values': numeric_values
    }

# --- Enhanced Query Processing ---
def enhance_query_with_intent(query, intent):
    """Enhance query based on detected intent."""
    query_lower = query.lower()
    
    if intent == "LOOKUP":
        # For lookups, focus on specific values and headers
        enhanced = query + " value numeric cell data"
        # Add context terms based on query
        if 'revenue' in query_lower:
            enhanced += " sales income"
        if 'target' in query_lower:
            enhanced += " goal budget"
        return enhanced
    
    elif intent == "CALCULATION":
        # For calculations, focus on aggregates and ranges
        enhanced = query + " sum total average aggregate range"
        return enhanced
    
    elif intent == "EXPLANATION":
        # For explanations, focus on formulas and relationships
        enhanced = query + " formula calculation dependency relationship"
        return enhanced
    
    # Default enhancement
    return query + " data value information"


# --- LLM Functions ---
def call_gemini(prompt, temperature=0.1):
    """Call Gemini model safely."""
    if not gemini_model_instance:
        return "LLM not available"
    
    try:
        response = gemini_model_instance.generate_content(
            prompt,
            generation_config=genai.types.GenerationConfig(
                temperature=temperature,
                max_output_tokens=2000
            )
        )
        return response.text.strip()
    except Exception as e:
        return f"Error: {str(e)}"
    
# --- Enhanced Search with Intent-Based Routing ---
def handle_lookup_intent(query, search_results):
    """Handle lookup queries - find specific values without calculations."""
    print("🔎 Handling LOOKUP intent")
    
    if not search_results['documents']:
        return "No matching data found."
    
    # Find the most relevant result for lookup
    best_result = None
    best_score = -1
    
    for i, (doc, meta, dist) in enumerate(zip(
        search_results['documents'],
        search_results['metadatas'],
        search_results['distances']
    )):
        confidence = 1 - (dist / 10) if dist < 10 else 0.1
        relevance = calculate_lookup_relevance(query, meta, confidence)
        
        if relevance > best_score:
            best_result = (meta, relevance)
            best_score = relevance
    
    if best_result and best_score > 0.5:  # Good match threshold
        meta, score = best_result
        return format_lookup_response(meta, query)
    
    # Fallback: show what was found without calculations
    return format_search_results(search_results, query, is_lookup=True)

def calculate_lookup_relevance(query, metadata, base_confidence):
    """Calculate relevance score for lookup queries."""
    query_lower = query.lower()
    relevance = base_confidence
    
    # Check header matches
    row_header = str(metadata.get('row_header', '')).lower()
    col_header = str(metadata.get('column_header', '')).lower()
    full_header = f"{row_header} {col_header}"
    
    # Boost for exact header matches
    if any(term in full_header for term in query_lower.split() if len(term) > 3):
        relevance += 0.3
    
    # Boost for numeric values (lookup queries often want specific numbers)
    if metadata.get('is_numeric', False):
        relevance += 0.2
    
    # Penalize formula cells for simple lookups (users usually want values)
    if metadata.get('formula'):
        relevance -= 0.1
    
    return min(relevance, 1.0)  # Cap at 1.0

def format_lookup_response(metadata, query):
    """Format response for lookup queries."""
    value = metadata.get('value', 'N/A')
    sheet = metadata.get('source_sheet', 'Unknown')
    cell = metadata.get('cell_address', '?')
    
    # Extract context from headers for better response
    row_header = metadata.get('row_header', '')
    col_header = metadata.get('column_header', '')
    
    if row_header and col_header:
        context = f"{row_header} - {col_header}"
    elif row_header:
        context = row_header
    elif col_header:
        context = col_header
    else:
        context = "the requested data"
    
    return f"The value for {context} is **{value}**. This was found in '{sheet}' at cell {cell}."

def format_search_results(results, query, is_lookup=False):
    """Format general search results."""
    if not results['documents']:
        return "No results found."
    
    output = []
    if is_lookup:
        output.append("🔍 Found these potential matches (showing top 5):")
    else:
        output.append("🔍 Search results (showing top 5):")
    
    for i, (doc, meta, dist) in enumerate(zip(
        results['documents'],
        results['metadatas'],
        results['distances']
    )):
        if i >= 5:
            break
        confidence = 1 - (dist / 10) if dist < 10 else 0.1
        output.append(f"{i+1}. {meta['source_sheet']} {meta['cell_address']} (conf: {confidence:.2f})")
        output.append(f"   Value: {meta.get('value', 'N/A')}")
        if meta.get('formula'):
            output.append(f"   Formula: {meta.get('formula')}")
        output.append(f"   Headers: {meta.get('row_header', '')} - {meta.get('column_header', '')}")
        output.append("")
    
    return "\n".join(output)

def handle_calculation_intent(query, search_results):
    """Handle calculation queries with proper filtering."""
    print("🧮 Handling CALCULATION intent")
    
    if not search_results['documents']:
        return "No data found for calculation."
    
    # Perform calculation with better filtering
    calculation_result = perform_calculations(search_results, query)
    
    if calculation_result:
        return format_calculation_response(calculation_result, query)
    
    return "Could not perform calculation on the found data."

def format_calculation_response(result, query):
    """Format calculation results."""
    output = []
    output.append(f"🧮 Calculation Results for: '{query}'")
    output.append(f"✅ Operation: {result['operation']}")
    output.append(f"📊 Result: {result['result']:.2f}")
    output.append(f"🔢 Based on: {result['count']} values")
    output.append("")
    output.append("📋 Values included in calculation:")
    
    for i, value in enumerate(result['values'][:10]):  # Show first 10 values
        output.append(f"  {i+1}. {value['sheet']} {value['cell']}: {value['value']}")
        if value.get('row_header') or value.get('column_header'):
            headers = f"{value.get('row_header', '')} - {value.get('column_header', '')}".strip(' -')
            if headers:
                output.append(f"     Context: {headers}")
    
    if len(result['values']) > 10:
        output.append(f"  ... and {len(result['values']) - 10} more values")
    
    return "\n".join(output)

def handle_explanation_intent(query, search_results):
    """Handle explanation queries - focus on formulas and relationships."""
    print("📚 Handling EXPLANATION intent")
    
    if not search_results['documents']:
        return "No formulas or explanatory content found."
    
    # Filter for formula cells
    formula_cells = []
    for meta in search_results['metadatas']:
        if meta.get('formula'):
            formula_cells.append(meta)
    
    if not formula_cells:
        return "No formulas found for explanation."
    
    output = []
    output.append(f"📚 Formula Analysis for: '{query}'")
    output.append(f"🔍 Found {len(formula_cells)} formulas:")
    
    for i, meta in enumerate(formula_cells[:5]):  # Show first 5 formulas
        output.append(f"\n{i+1}. {meta['source_sheet']} {meta['cell_address']}")
        output.append(f"   Formula: {meta.get('formula')}")
        output.append(f"   Value: {meta.get('value', 'N/A')}")
        output.append(f"   Context: {meta.get('row_header', '')} - {meta.get('column_header', '')}")
        if meta.get('dependencies'):
            output.append(f"   Dependencies: {', '.join(meta['dependencies'][:3])}")
            if len(meta['dependencies']) > 3:
                output.append(f"     ... and {len(meta['dependencies']) - 3} more")
    
    return "\n".join(output)

def handle_comparison_intent(query, search_results):
    """Handle comparison queries."""
    print("⚖️ Handling COMPARISON intent")
    
    # For now, use general search results format
    return format_search_results(search_results, query, is_lookup=False)

def first_llm_call_translator(user_query, search_results):
    """First LLM call for query enhancement with semantic understanding."""
    context_text = "\n".join([
        f"From {meta['source_sheet']} {meta['cell_address']}: {doc[:150]}..."
        for doc, meta in zip(search_results['documents'][:3], search_results['metadatas'][:3])
    ])
    
    # First detect intent to guide the enhancement
    intent = detect_query_intent(user_query)
    
    prompt = f"""
    ROLE: You are a data analyst enhancing search queries for spreadsheet data.

    USER QUERY: "{user_query}"
    DETECTED INTENT: {intent}

    SAMPLE DATA FOUND:
    {context_text}

    TASK: Create a semantically enhanced search query based on the detected intent.

    {f"**LOOKUP INTENT**: Focus on finding specific values, headers, and exact matches." if intent == "LOOKUP" else ""}
    {f"**CALCULATION INTENT**: Focus on finding numeric values for computation. Include specific operations like SUM, AVERAGE, etc." if intent == "CALCULATION" else ""}
    {f"**EXPLANATION INTENT**: Focus on finding formulas, cell relationships, and explanatory content. Prioritize formula cells and their dependencies." if intent == "EXPLANATION" else ""}
    {f"**COMPARISON INTENT**: Focus on finding comparable values, metrics, and related data points." if intent == "COMPARISON" else ""}

    OUTPUT: Only the enhanced search query:
    """
    
    return call_gemini(prompt, temperature=0.1)

def second_llm_call_analyst(user_query: str, search_results: Dict, calculation_result: Dict = None, intent: str = "LOOKUP") -> str:
    """
    Enhanced LLM call with semantic intent understanding.
    """
    # Format context differently based on intent
    if intent in ["EXPLANATION"]:
        # For explanatory queries, focus on formulas and metadata
        context_text = "\n".join([
            f"📋 {meta.get('source_sheet', 'Unknown')} {meta.get('cell_address', '?')} | "
            f"Formula: {meta.get('formula', 'None')} | "
            f"Value: {meta.get('value', 'N/A')} | "
            f"Headers: {meta.get('row_header', '')} - {meta.get('column_header', '')}"
            for _, meta in zip(search_results.get('documents', []), search_results.get('metadatas', []))
        ])
    else:
        # For other intents, use the existing format
        context_text = "\n".join([
            f"📊 {meta.get('source_sheet', 'Unknown')} {meta.get('cell_address', '?')} | "
            f"Row: {meta.get('row_header', '')} | Column: {meta.get('column_header', '')} | "
            f"Value: {meta.get('value', 'N/A')}"
            for _, meta in zip(search_results.get('documents', []), search_results.get('metadatas', []))
        ])
    
    # Add calculation results to prompt if available
    calculation_context = ""
    if calculation_result:
        value_details = "\n".join([
            f"  - {item['sheet']} {item['cell']}: {item['value']} ({item.get('header', '')})"
            for item in calculation_result['values'][:5]
        ])
        
        calculation_context = f"""
        CALCULATION RESULTS:
        Operation: {calculation_result['operation']}
        Result: {calculation_result['result']:.2f}
        Values Used: {len(calculation_result['values'])} numeric values
        
        Values Included:
        {value_details}
        """

    # Create intent-aware prompt
    if intent == "CALCULATION":
        prompt = f"""
        ROLE: You are a data analyst providing calculation results.

        USER QUERY: {user_query}
        QUERY INTENT: {intent}

        RELEVANT DATA FOUND:
        {context_text}

        {calculation_context}

        TASK: Present the calculation results clearly and explain what they represent.
        Focus on the business meaning and context of the results.
        """
    
    elif intent == "EXPLANATION":
        prompt = f"""
        ROLE: You are a data analyst providing information about spreadsheet formulas and data.

        USER QUERY: {user_query}
        QUERY INTENT: {intent}

        RELEVANT FORMULAS AND DATA FOUND:
        {context_text}

        TASK: Explain the formulas and data relationships found. Focus on:
        1. What each formula calculates and its business purpose
        2. How cells are related through dependencies
        3. The context and meaning of the formulas
        4. Any insights or patterns discovered
        """
    
    elif intent == "LOOKUP":
        prompt = f"""
        ROLE: You are a data analyst helping locate specific spreadsheet data.

        USER QUERY: {user_query}
        QUERY INTENT: {intent}

        RELEVANT DATA FOUND:
        {context_text}

        TASK: Help the user understand what specific data was found and where it's located.
        Provide clear context about the meaning and significance of the found data.
        """
    
    else:  # COMPARISON and other intents
        prompt = f"""
        ROLE: You are a data analyst providing comprehensive search results.

        USER QUERY: {user_query}
        QUERY INTENT: {intent}

        RELEVANT DATA FOUND:
        {context_text}

        TASK: Provide a comprehensive overview of what data was found.
        Help the user understand the context, relationships, and significance of the results.
        """

    return call_gemini(prompt, temperature=0.0)

# --- Vector Database Functions ---
def populate_vector_database(chunks):
    """Populate FAISS index."""
    global stored_texts, stored_metadata
    
    texts = [chunk["text"] for chunk in chunks]
    print(f"Processing {len(texts)} chunks...")
    
    # Generate embeddings
    embeddings = embedder.encode(texts)
    
    # Store data
    stored_texts = texts
    stored_metadata = [chunk["metadata"] for chunk in chunks]
    
    # Add to FAISS
    embeddings_array = np.array(embeddings).astype('float32')
    index.add(embeddings_array)
    
    print(f"✅ Added {len(stored_texts)} documents to index")
    return True


def faiss_search(query, n_results=100):
    """Search using FAISS."""
    query_embedding = embedder.encode([query]).astype('float32')
    distances, indices = index.search(query_embedding, n_results)
    
    results = {
        'documents': [],
        'metadatas': [],
        'distances': distances[0].tolist()
    }
    
    for idx in indices[0]:
        if 0 <= idx < len(stored_texts):
            results['documents'].append(stored_texts[idx])
            results['metadatas'].append(stored_metadata[idx])
    
    return results

# --- Main Semantic Search Function ---
def semantic_search(query):
    """Main search function with enhanced intent-based routing."""
    print(f"\n🔍 Searching: '{query}'")
    
    try:
        # Detect intent first
        intent = detect_query_intent(query)
        print(f"🎯 Detected intent: {intent}")
        
        # Adjust search parameters based on intent
        if intent == "CALCULATION":
            n_results = 100  # Need more results for calculations
        elif intent == "EXPLANATION":
            n_results = 50   # Focus on quality for explanations
        elif intent == "LOOKUP":
            n_results = 75   # Balanced approach for lookups
        else:
            n_results = 60   # Default for other intents
        
        # Enhance query based on intent
        enhanced_query = enhance_query_with_intent(query, intent)
        print(f"🔍 Enhanced query: {enhanced_query}")
        
        # Search with intent-appropriate parameters
        results = faiss_search(enhanced_query, n_results)
        if not results['documents']:
            return "No data found. Try different terms."
        
        # Show what search found
        print(f"\n📋 Search found {len(results['documents'])} results")
        print(f"Top {min(10, n_results//3)} results:")
        
        for i, (doc, meta, dist) in enumerate(zip(
            results['documents'],
            results['metadatas'],
            results['distances']
        )):
            if i >= min(10, n_results//3):
                break
            confidence = 1 - (dist / 10) if dist < 10 else 0.1
            print(f"{i+1}. {meta['source_sheet']} {meta['cell_address']} (conf: {confidence:.2f})")
            if intent == "EXPLANATION" and meta.get('formula'):
                print(f"   Formula: {meta.get('formula', 'None')}")
            print(f"   Value: {meta.get('value', 'N/A')}")
            print(f"   Headers: {meta.get('row_header', '')} - {meta.get('column_header', '')}")
        
        # Initialize calculation_result variable
        calculation_result = None
        
        # Route to appropriate handler based on intent and get response
        if intent == "LOOKUP":
            response = handle_lookup_intent(query, results)
        elif intent == "CALCULATION":
            calculation_result = perform_calculations(results, query)
            response = handle_calculation_intent(query, results)
        elif intent == "EXPLANATION":
            response = handle_explanation_intent(query, results)
        elif intent == "COMPARISON":
            response = handle_comparison_intent(query, results)
        else:
            response = format_search_results(results, query, is_lookup=False)
        
        # Generate final LLM-enhanced response
        final_response = second_llm_call_analyst(query, results, calculation_result, intent)
        
        return final_response
        
    except Exception as e:
        return f"Search error: {str(e)}"
    
# --- Debug and Main Execution Functions ---
def debug_chunk_creation(chunks, sample_size=10):
    """Debug function to see what chunks were created."""
    print(f"\n🔍 DEBUG: Sample of {sample_size} chunks created:")
    for i, chunk in enumerate(chunks[:sample_size]):
        print(f"{i+1}. Sheet: {chunk['metadata']['source_sheet']}")
        print(f"   Cell: {chunk['metadata']['cell_address']}")
        print(f"   Value: {chunk['metadata']['value']}")
        print(f"   Headers: {chunk['metadata']['row_header']} - {chunk['metadata']['column_header']}")
        print(f"   Text: {chunk['text'][:100]}...")
        print()

# Update the semantic_search call in main execution
if __name__ == "__main__":
    try:
        file_path = "Sales Dashboard.xlsx"
        print("🚀 Starting Structured Semantic Search...")
        
        # Parse with proper table detection
        chunks, metadata, numeric_data = parse_spreadsheet_structured(file_path)
        
        # Debug: show what chunks were created
        debug_chunk_creation(chunks)
        
        # Populate database
        success = populate_vector_database(chunks)
        
        if success:
            print(f"\n✅ Ready! Loaded {len(stored_texts)} cells")
            print(f"✅ Numeric data: {len(global_numeric_data)} values available for calculations")
            while True:
                query = input("\n🔍 Your question (or 'exit'): ").strip()
                if query.lower() == 'exit':
                    break
                if not query:
                    continue
                
                result = semantic_search(query)
                print(f"\n💡 Answer: {result}")
                print("─" * 60)
                
    except Exception as e:
        print(f"❌ Error: {str(e)}")
